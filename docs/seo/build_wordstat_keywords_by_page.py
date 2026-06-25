#!/usr/bin/env python3
"""Собирает wordstat-keywords-by-page.xlsx из выгрузки Вордстата.

Правила:
- ключи распределяются по наиболее подходящей странице;
- страница «Каталогизация (раздел)» получает общие запросы + топ-N с подстраниц раздела;
- хабы «Статьи» и «Учебный центр» — аналогично с дочерними страницами.

Источник по умолчанию: ~/Desktop/WORD STAT.xlsx, лист «Лист1».
"""

from __future__ import annotations

import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[2]
OUT = Path(__file__).resolve().parent / "wordstat-keywords-by-page.xlsx"
DEFAULT_SRC = Path.home() / "Desktop" / "WORD STAT.xlsx"

TOP_PER_CHILD = 10

# Фразы с текстов страниц / узкой тематики — добавляются, если нет в выгрузке Вордстата.
SUPPLEMENT: dict[str, list[str]] = {
    "KI": [
        "каталогизация комплектующих изделий",
        "каталогизация комплектующих изделий и материалов",
        "каталогизация материалов военного назначения",
        "каталогизация ки и мвн",
        "приказ минпромторга 4725 каталогизация",
        "каталогизация по приказу 4725",
    ],
    "SFO": [
        "разработка стандартных форматов описания",
        "разработка сфо",
        "разработка сфо пс",
        "гост рв 0044-006-2025",
        "стандартный формат описания предметов снабжения сфо",
    ],
    "FKP": [
        "проверка наличия продукции в фкп",
        "проверка наличия в федеральном каталоге продукции",
        "уведомление о наличии продукции в фкп",
        "уведомление о наличии в фкп",
        "наличие продукции в федеральном каталоге",
    ],
}

# ВЧ-запросы для отдельного прогона в Вордстате (если нужно ещё расширить ядро).
WORDSTAT_SEEDS: dict[str, list[str]] = {
    "KI": [
        "каталогизация комплектующих изделий",
        "каталогизация мвн",
        "приказ 4725 каталогизация",
        "каталогизация предметов снабжения",
    ],
    "SFO": [
        "разработка сфо",
        "сфо предметов снабжения",
        "гост рв 0044 006",
    ],
    "FKP": [
        "проверка наличия в фкп",
        "уведомление о наличии фкп",
        "федеральный каталог продукции проверка",
    ],
}

LABEL = {
    "HUB": "Каталогизация (раздел /catalogization)",
    "GOZ": "Каталогизация продукции для федеральных нужд",
    "KI": "Каталогизация продукции КИ и МВН",
    "SFO": "Разработка стандартных форматов описания (СФО)",
    "FKP": "Проверка наличия продукции в ФКП",
    "TRAIN_HUB": "Учебный центр (раздел /training-center)",
    "TRAIN_LIT": "Учебная литература по каталогизации",
    "ART_HUB": "Статьи (раздел /articles)",
    "ART_WHAT": "Зачем нужна каталогизация продукции",
    "ART_UNIF": "Техрегулирование в ОПК",
}

ORDER = list(LABEL.keys())

HUB_CHILDREN: dict[str, list[str]] = {
    "HUB": ["GOZ", "KI", "SFO", "FKP"],
    "ART_HUB": ["ART_WHAT", "ART_UNIF"],
    "TRAIN_HUB": ["TRAIN_LIT"],
}

EXCLUDE = re.compile(
    r"библиотек|книг[аиуе]?|литератур(?!.*каталог)|домашн|музе|архивн|"
    r"фото(?!граф)|фотограф|изображен|видео|фильм|музык|"
    r"беларус|белорус|белгисс|\bрб\b|"
    r"каталогизаци[ия] сайт|приложени[ея]|"
    r"ресурсов образован|каис|"
    r"метролог|\bии каталогизац|сукиасян|"
    r"код ун\b|машиночитаем|"
    r"отдел каталогизации|библиотечн|комплектование каталогизац|"
    r"коллекци[ия]|каталогизаци[ия] файлов|"
    r"каталогизаци[ия] книг|"
    r"унификация и каталогизация экспертных|проблемы алгоритмизации|"
    r"7703211463|где взять номер сфо|"
    r"корпоративные каталогизации|"
    r"определение комплектующ|определение по гост|примеры комплектующ|"
    r"специалист по каталогизац|"
    r"скачать|\.pdf\b",
    re.I,
)

LAW_ONLY = re.compile(
    r"^(?:275\s*фз|закон\s+275|федеральн[ыйого]+\s+закон\s+275|"
    r"статья\s+\d|ст\s+275|гособоронзаказ\s+статья|"
    r"\d+\.?\d*\s+275\s*фз|275\s*фз\s+\d|275\s*фз\s+от\s+29)",
    re.I,
)


def classify(q: str) -> Optional[str]:
    ql = q.casefold()
    has_cat = "каталогиз" in ql

    if EXCLUDE.search(q):
        return None
    if LAW_ONLY.search(q.strip()) and not has_cat:
        return None
    if re.search(r"\bстатья\s+\d|\bст\s+\d", ql) and "275" in ql and not has_cat:
        return None
    if re.search(r"275\s*фз", ql) and not has_cat and not re.search(r"каталог|гоз|оборон", ql):
        return None

    if re.search(r"моисеев", ql):
        return "TRAIN_LIT"
    if re.search(r"обучени[ея].*каталог|каталог.*обучени[ея]|курс[ыа]?.*каталог|каталог.*курс", ql):
        return "TRAIN_HUB"

    if has_cat and re.search(
        r"что такое|зачем|это простыми|определен|вопросы каталогизац|"
        r"этапы каталогизац|процесс каталогизац|классификация и каталогизац",
        ql,
    ):
        return "ART_WHAT" if re.search(r"что такое|зачем|простыми|определен", ql) else "ART_HUB"
    if has_cat and re.search(r"стандартизац|унификац", ql):
        return "ART_UNIF"

    if re.search(
        r"проверк.*налич|наличи[ея].*фкп|уведомлени[ея].*налич|\bфкп\b|"
        r"федеральн[ыйого]+\s+каталог\s+продукц|федеральн[ыйого]+\s+каталог\s+предметов",
        ql,
    ):
        return "FKP"
    if re.search(r"\bсфо\b|стандартн[ыйого]+\s+формат|код\s+сфо", ql):
        return "SFO"
    if re.search(r"гост\s*рв\s*0044", ql):
        # СФО по ГОСТ РВ 0044-006; общий запрос «гост рв 0044» тоже уместен
        if re.search(r"0044\s*006", ql) or re.fullmatch(r"гост\s*рв\s*0044", ql.strip()):
            return "SFO"
    if re.search(r"\b4725\b|приказ\s+минпромторга", ql):
        return "KI"
    if has_cat and re.search(
        r"предмет.*снабжен|комплектующ|"
        r"\b4725\b|минпромторг.*каталог|каталог.*комплектующ|"
        r"\bмвн\b|\bки\b.*каталог|каталог.*\bки\b|вс\s*рф|мо\s*рф",
        ql,
    ):
        return "KI"
    if has_cat and re.search(
        r"\bгоз\b|гособорон|275\s*фз|549|фнн|федеральн|оборонн|подлежащ|"
        r"изделий|военн|государственн.*нужд|работ.*каталог|выполнен.*каталог|"
        r"проведен.*каталог|номенклатур|каталожн|под ключ",
        ql,
    ):
        return "GOZ"
    if has_cat:
        return "HUB"
    return None


def parse_freq(value) -> int:
    if value is None:
        return 0
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def load_queries(src: Path) -> list[tuple[str, int]]:
    wb = load_workbook(src, data_only=True)
    ws = wb["Лист1"]
    best: dict[str, tuple[str, int]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        q, freq = row[0], row[1]
        if not q:
            continue
        q = str(q).strip()
        f = parse_freq(freq)
        key = q.casefold()
        if key not in best or best[key][1] < f:
            best[key] = (q, f)
    return list(best.values())


def build_rows(queries: list[tuple[str, int]]) -> list[tuple[str, str, int]]:
    freq_by_kw = {q.casefold(): (q, f) for q, f in queries}

    by_page: dict[str, list[tuple[str, int]]] = defaultdict(list)
    for q, f in queries:
        code = classify(q)
        if code:
            by_page[code].append((q, f))

    for code, phrases in SUPPLEMENT.items():
        seen = {q.casefold() for q, _ in by_page.get(code, [])}
        for phrase in phrases:
            key = phrase.casefold()
            if key in seen:
                continue
            seen.add(key)
            q, f = freq_by_kw.get(key, (phrase, 0))
            by_page[code].append((q, f))

    for code in by_page:
        by_page[code].sort(key=lambda x: (-x[1], x[0]))

    # хабы: добавляем топы с дочерних страниц
    for hub, children in HUB_CHILDREN.items():
        seen = {q.casefold() for q, _ in by_page.get(hub, [])}
        for child in children:
            for q, f in by_page.get(child, [])[:TOP_PER_CHILD]:
                if q.casefold() not in seen:
                    by_page[hub].append((q, f))
                    seen.add(q.casefold())
        by_page[hub].sort(key=lambda x: (-x[1], x[0]))

    rows: list[tuple[str, str, int]] = []
    for code in ORDER:
        for q, f in by_page.get(code, []):
            rows.append((LABEL[code], q, f))
    return rows


def build_summary(rows: list[tuple[str, str, int]]) -> list[tuple[str, str]]:
    """Сводка: раздел → ключи через запятую (по убыванию частотности)."""
    by_section: dict[str, list[tuple[int, str]]] = defaultdict(list)
    for section, q, f in rows:
        by_section[section].append((f, q))

    summary: list[tuple[str, str]] = []
    seen_sections: set[str] = set()
    for code in ORDER:
        section = LABEL[code]
        if section in seen_sections or section not in by_section:
            continue
        seen_sections.add(section)
        items = sorted(by_section[section], key=lambda x: (-x[0], x[1]))
        keywords = ", ".join(q for _, q in items)
        summary.append((section, keywords))
    return summary


def _style_header(ws, cols: int) -> None:
    for col in range(1, cols + 1):
        cell = ws.cell(1, col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="496DB3")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def write_xlsx(rows: list[tuple[str, str, int]], path: Path) -> None:
    summary = build_summary(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "Ключи по страницам"
    ws.append(["Раздел / страница", "Ключ", "Частотность"])
    _style_header(ws, 3)
    for section, q, f in rows:
        ws.append([section, q, f])
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 56
    ws.column_dimensions["C"].width = 14

    ws_sum = wb.create_sheet("Сводка по разделам", 0)
    ws_sum.append(["Раздел / страница", "Ключи"])
    _style_header(ws_sum, 2)
    for section, keywords in summary:
        ws_sum.append([section, keywords])
        ws_sum.cell(ws_sum.max_row, 2).alignment = Alignment(wrap_text=True, vertical="top")
    ws_sum.column_dimensions["A"].width = 52
    ws_sum.column_dimensions["B"].width = 96

    ws_seeds = wb.create_sheet("Семена для Вордстата")
    ws_seeds.append(["Раздел / страница", "Запрос для прогона"])
    _style_header(ws_seeds, 2)
    for code in ORDER:
        seeds = WORDSTAT_SEEDS.get(code)
        if not seeds:
            continue
        for seed in seeds:
            ws_seeds.append([LABEL[code], seed])
    ws_seeds.column_dimensions["A"].width = 52
    ws_seeds.column_dimensions["B"].width = 56

    wb.save(path)


def sync_desktop_sheets(src: Path, rows: list[tuple[str, str, int]]) -> None:
    if not src.exists():
        return
    wb = load_workbook(src)
    summary = build_summary(rows)
    for name in ["запросы вордстат", "Ключи по страницам", "Сводка по разделам"]:
        if name in wb.sheetnames:
            del wb[name]
    ws_sum = wb.create_sheet("Сводка по разделам", 0)
    ws_sum.append(["Раздел / страница", "Ключи"])
    for section, keywords in summary:
        ws_sum.append([section, keywords])
    ws_z = wb.create_sheet("запросы вордстат")
    ws_z.append(["Услуга", "Запрос", "Частотность"])
    for section, q, f in rows:
        ws_z.append([section, q, f])
    ws_k = wb.create_sheet("Ключи по страницам")
    ws_k.append(["Раздел / страница", "Ключ", "Частотность"])
    for section, q, f in rows:
        ws_k.append([section, q, f])
    wb.save(src)


def main() -> None:
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_SRC
    if not src.exists():
        raise SystemExit(f"Нет файла: {src}")

    queries = load_queries(src)
    rows = build_rows(queries)
    write_xlsx(rows, OUT)
    sync_desktop_sheets(src, rows)

    hub = [r for r in rows if r[0] == LABEL["HUB"]]
    print(f"Записано: {OUT}")
    print(f"Всего строк: {len(rows)}")
    print(f"Раздел «Каталогизация»: {len(hub)} ключей (топ-{TOP_PER_CHILD} с каждой подстраницы включены)")
    print("Топ-15 раздела:")
    for _, q, f in sorted(hub, key=lambda x: -x[2])[:15]:
        print(f"  {f:>5}  {q}")


if __name__ == "__main__":
    main()

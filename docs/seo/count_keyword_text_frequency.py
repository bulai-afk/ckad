#!/usr/bin/env python3
"""Считает вхождения ключевых фраз в тексте страниц каталогизации.

Заполняет колонку «Частотность (в тексте)» в catalogization-keywords.xlsx.
Не использует API Вордстата.

  python3 docs/seo/count_keyword_text_frequency.py
"""

from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
XLSX_PATH = Path(__file__).resolve().parent / "catalogization-keywords.xlsx"
SHEET_NAME = "Семантическое ядро"
FREQ_HEADER = "Частотность (в тексте)"

SERVICE_TO_SLUG = {
    "Каталогизация (раздел /catalogization)": "__hub__",
    "Каталогизация продукции для федеральных нужд": "catalogization/katalogizatsiya-produktsii-po-goz-pod-klyuch",
    "Каталогизация продукции КИ и МВН": "catalogization/katalogizatsiya-ki-i-mvn",
    "Разработка стандартных форматов описания (СФО)": "catalogization/razrabotka-standartnyh-formatov-opisaniya-sfo",
    "Проверка наличия продукции в ФКП": "catalogization/proverka-nalichiya-produktsii-v-fkp",
}

HOME_SERVICES_INTRO = (
    "Помогаем участникам ГОЗ выполнять требования государственных контрактов в части "
    "каталогизации, применения продукции иностранного производства, разработки электронной "
    "конструкторской документации, поиска взаимозаменяемых аналогов и другие."
)

NOISE = re.compile(
    r"Мониторинг серверов|Совместная работа|Планирование задач|Подробнее|Добавьте изображение",
    re.I,
)


def strip_html(html: str) -> str:
    html = re.sub(r"<script[^>]*>[\s\S]*?</script>", " ", html or "", flags=re.I)
    html = re.sub(r"<style[^>]*>[\s\S]*?</style>", " ", html, flags=re.I)
    html = re.sub(r"<[^>]+>", " ", html)
    html = html.replace("&nbsp;", " ")
    return re.sub(r"\s+", " ", html).strip()


def page_text(page: dict) -> str:
    parts = [page.get("title") or "", page.get("description") or ""]
    for block in page.get("blocks") or []:
        data = block.get("data") or {}
        raw = data.get("text")
        if not raw:
            continue
        parts.append(strip_html(raw) if block.get("type") == "text" else str(raw))
    return NOISE.sub(" ", " ".join(parts))


def build_corpus() -> dict[str, str]:
    pages = json.loads((ROOT / "backend/data/pages.import.json").read_text(encoding="utf-8"))
    folders = json.loads((ROOT / "backend/data/folders.json").read_text(encoding="utf-8"))
    cat_folder = next((f for f in folders if f.get("slug") == "catalogization"), {})

    corpus: dict[str, str] = {}
    for service_name, slug in SERVICE_TO_SLUG.items():
        if slug == "__hub__":
            parts = [
                cat_folder.get("name") or "",
                cat_folder.get("description") or "",
                HOME_SERVICES_INTRO,
            ]
            for page in pages:
                if page.get("slug", "").startswith("catalogization/") and page.get("status") == "PUBLISHED":
                    parts.append(page_text(page))
            corpus[service_name] = NOISE.sub(" ", " ".join(parts))
            continue

        page = next((p for p in pages if p.get("slug") == slug), None)
        corpus[service_name] = page_text(page) if page else ""

    return corpus


def count_keyword(text: str, keyword: str) -> int:
    if not text or not keyword:
        return 0
    return text.casefold().count(keyword.casefold())


def ensure_freq_column(ws) -> int:
    for col in range(1, ws.max_column + 1):
        header = ws.cell(1, col).value
        if header in ("Частотность", FREQ_HEADER):
            ws.cell(1, col, FREQ_HEADER)
            return col
    col = ws.max_column + 1
    ws.cell(1, col, FREQ_HEADER)
    return col


def main() -> None:
    corpus = build_corpus()
    wb = load_workbook(XLSX_PATH)
    ws = wb[SHEET_NAME]
    freq_col = ensure_freq_column(ws)

    for row in range(2, ws.max_row + 1):
        service = str(ws.cell(row, 1).value or "").strip()
        keyword = str(ws.cell(row, 2).value or "").strip()
        ws.cell(row, freq_col, count_keyword(corpus.get(service, ""), keyword))

    wb.save(XLSX_PATH)
    print(f"Готово: {XLSX_PATH}")


if __name__ == "__main__":
    main()

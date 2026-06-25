#!/usr/bin/env python3
"""Подтягивает частотность ключей из Yandex Wordstat API v2 в catalogization-keywords.xlsx.

Требуется Yandex Cloud Search API (AI Studio):
  YANDEX_WORDSTAT_API_KEY  — API-ключ сервисного аккаунта (или YANDEX_SEARCH_API_KEY)
  YANDEX_WORDSTAT_FOLDER_ID — ID каталога (или YANDEX_FOLDER_ID)

Запуск из корня репозитория:
  python3 docs/seo/fetch_wordstat_frequencies.py
  python3 docs/seo/fetch_wordstat_frequencies.py --dry-run
"""

from __future__ import annotations

import argparse
import json
import os
import re
import ssl
import sys
import time
import urllib.error
import urllib.request
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[2]
XLSX_PATH = Path(__file__).resolve().parent / "catalogization-keywords.xlsx"
SHEET_NAME = "Семантическое ядро"
API_BASE = "https://searchapi.api.cloud.yandex.net/v2/wordstat"

# Москва + область — типичный B2B-контекст; пустой список = вся Россия в UI Вордстата
DEFAULT_REGIONS: list[str] | None = None

SSL_CTX = ssl.create_default_context()
if os.environ.get("WORDSTAT_SSL_VERIFY", "1") == "0":
    SSL_CTX.check_hostname = False
    SSL_CTX.verify_mode = ssl.CERT_NONE


def normalize_phrase(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip().casefold())


def get_credentials(folder_id_arg: str | None = None) -> tuple[str, str]:
    api_key = (
        os.environ.get("YANDEX_WORDSTAT_API_KEY", "").strip()
        or os.environ.get("YANDEX_SEARCH_API_KEY", "").strip()
    )
    folder_id = (
        (folder_id_arg or "").strip()
        or os.environ.get("YANDEX_WORDSTAT_FOLDER_ID", "").strip()
        or os.environ.get("YANDEX_FOLDER_ID", "").strip()
    )
    if api_key.startswith("aje") and len(api_key) < 30:
        raise SystemExit(
            "Похоже, вы указали идентификатор ключа (aje...), а не секретный ключ.\n"
            "Нужна длинная строка, которая показывается один раз при создании (обычно начинается с AQVN...).\n"
            "Создайте новый API-ключ в AI Studio, если старый секрет не сохранили."
        )
    if not api_key or not folder_id:
        raise SystemExit(
            "Не заданы YANDEX_WORDSTAT_API_KEY и YANDEX_WORDSTAT_FOLDER_ID "
            "(или YANDEX_SEARCH_API_KEY / YANDEX_FOLDER_ID).\n"
            "Инструкция — лист «Справка» в catalogization-keywords.xlsx"
        )
    return api_key, folder_id


def wordstat_call(api_key: str, path: str, body: dict) -> dict:
    payload = json.dumps(body, ensure_ascii=False).encode("utf-8")
    req = urllib.request.Request(
        f"{API_BASE}/{path}",
        data=payload,
        headers={
            "Authorization": f"Api-Key {api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=45, context=SSL_CTX) as resp:
            return json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        body_text = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"HTTP {exc.code}: {body_text[:500]}") from exc


def fetch_frequency(api_key: str, folder_id: str, phrase: str) -> tuple[int | None, str]:
    body: dict = {
        "phrase": phrase,
        "numPhrases": 50,
        "folderId": folder_id,
    }
    if DEFAULT_REGIONS:
        body["regions"] = DEFAULT_REGIONS

    data = wordstat_call(api_key, "topRequests", body)
    if "code" in data and "message" in data:
        raise RuntimeError(data.get("message", str(data)))

    target = normalize_phrase(phrase)
    for item in data.get("results", []):
        if normalize_phrase(str(item.get("phrase", ""))) == target:
            return int(item["count"]), "точное совпадение"

    total = data.get("totalCount")
    if total is not None and str(total).isdigit():
        return int(total), "totalCount (широкое вхождение)"

    return None, "нет данных"


def ensure_frequency_column(ws) -> int:
    headers = [cell.value for cell in ws[1]]
    if "Частотность" in headers:
        return headers.index("Частотность") + 1
    col = len(headers) + 1
    cell = ws.cell(row=1, column=col, value="Частотность")
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="496DB3")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.column_dimensions[chr(64 + col) if col <= 26 else "C"].width = 18
    return col


def update_help_sheet(wb) -> None:
    if "Справка" not in wb.sheetnames:
        return
    ws = wb["Справка"]
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
        if row[0].value == "Подключение Wordstat API":
            return
    rows = [
        [],
        ["Подключение Wordstat API", ""],
        [
            "1",
            "Yandex Cloud → включить Search API / AI Studio (search-api.executor на сервисном аккаунте)",
        ],
        ["2", "Создать API-ключ сервисного аккаунта"],
        ["3", "Скопировать folder ID из консоли Yandex Cloud (каталог)"],
        [
            "4",
            "export YANDEX_WORDSTAT_API_KEY='...' && export YANDEX_WORDSTAT_FOLDER_ID='b1g...'",
        ],
        ["5", "python3 docs/seo/fetch_wordstat_frequencies.py"],
        [],
        ["Частотность", "показов в месяц, Россия (без фильтра регионов)"],
        [
            "Метод",
            "POST /v2/wordstat/topRequests — сначала точное совпадение фразы, иначе totalCount",
        ],
        ["Пауза между запросами", "0.25 с (снижение риска 429)"],
    ]
    start = ws.max_row + 1
    for offset, row in enumerate(rows):
        for col, value in enumerate(row, start=1):
            ws.cell(row=start + offset, column=col, value=value)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true", help="Только показать ключи, без API и записи")
    parser.add_argument(
        "--init-only",
        action="store_true",
        help="Добавить колонку «Частотность» и инструкцию, без запросов к API",
    )
    parser.add_argument("--sleep", type=float, default=0.25, help="Пауза между запросами, сек")
    parser.add_argument("--folder-id", help="ID каталога Yandex Cloud (b1g...)")
    args = parser.parse_args()

    if not XLSX_PATH.exists():
        raise SystemExit(f"Файл не найден: {XLSX_PATH}")

    wb = load_workbook(XLSX_PATH)
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"Лист «{SHEET_NAME}» не найден")

    ws = wb[SHEET_NAME]
    freq_col = ensure_frequency_column(ws)

    keywords: list[str] = []
    for row in range(2, ws.max_row + 1):
        kw = ws.cell(row=row, column=2).value
        if kw:
            keywords.append(str(kw).strip())

    unique = list(dict.fromkeys(keywords))
    print(f"Ключей в таблице: {len(keywords)}, уникальных: {len(unique)}")

    if args.dry_run:
        for kw in unique[:10]:
            print(f"  - {kw}")
        if len(unique) > 10:
            print(f"  ... и ещё {len(unique) - 10}")
        return

    if args.init_only:
        update_help_sheet(wb)
        wb.save(XLSX_PATH)
        print(f"Колонка «Частотность» добавлена: {XLSX_PATH}")
        return

    api_key, folder_id = get_credentials(args.folder_id)
    cache: dict[str, tuple[int | None, str]] = {}

    for index, phrase in enumerate(unique, start=1):
        if phrase in cache:
            continue
        print(f"[{index}/{len(unique)}] {phrase}")
        try:
            cache[phrase] = fetch_frequency(api_key, folder_id, phrase)
            count, kind = cache[phrase]
            print(f"    → {count if count is not None else '—'} ({kind})")
        except RuntimeError as exc:
            print(f"    Ошибка: {exc}", file=sys.stderr)
            cache[phrase] = (None, f"ошибка: {exc}")
            if "HTTP 401" in str(exc) or "HTTP 403" in str(exc):
                raise SystemExit(
                    "\nОстановка: проверьте секретный ключ (AQVN...) и доступ Search API."
                ) from exc
        time.sleep(args.sleep)

    for row in range(2, ws.max_row + 1):
        phrase = str(ws.cell(row=row, column=2).value or "").strip()
        count, _kind = cache.get(phrase, (None, ""))
        cell = ws.cell(row=row, column=freq_col)
        cell.value = count if count is not None else None
        cell.alignment = Alignment(vertical="top")

    update_help_sheet(wb)
    wb.save(XLSX_PATH)
    filled = sum(1 for _, (c, _) in cache.items() if c is not None)
    print(f"\nГотово: {XLSX_PATH}")
    print(f"Заполнено частотностей: {filled}/{len(unique)}")


if __name__ == "__main__":
    main()

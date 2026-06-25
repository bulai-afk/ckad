"""Общие константы и чтение catalogization-keywords.xlsx."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
XLSX_PATH = Path(__file__).resolve().parent / "catalogization-keywords.xlsx"
SHEET_CORE = "Семантическое ядро"
SHEET_WORDSTAT = "Вордстат — ввод"
SHEET_CMS = "Keywords для CMS"

SERVICE_TO_SLUG: dict[str, str] = {
    "Каталогизация (раздел /catalogization)": "__hub__",
    "Каталогизация продукции для федеральных нужд": "catalogization/katalogizatsiya-produktsii-po-goz-pod-klyuch",
    "Каталогизация продукции КИ и МВН": "catalogization/katalogizatsiya-ki-i-mvn",
    "Разработка стандартных форматов описания (СФО)": "catalogization/razrabotka-standartnyh-formatov-opisaniya-sfo",
    "Проверка наличия продукции в ФКП": "catalogization/proverka-nalichiya-produktsii-v-fkp",
}

SLUG_TO_SERVICE = {v: k for k, v in SERVICE_TO_SLUG.items() if v != "__hub__"}

HEADER_SERVICE = "Услуга"
HEADER_KEYWORD = "Ключевое слово"
HEADER_TEXT_FREQ = "Частотность (в тексте)"
HEADER_WS_FREQ = "Частотность (Вордстат)"
HEADER_IN_META = "В meta"


def header_map(ws) -> dict[str, int]:
    return {str(c.value): idx + 1 for idx, c in enumerate(ws[1]) if c.value}


def parse_int(value) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return int(value)
    s = str(value).strip().replace(" ", "").replace("\u00a0", "")
    if not s:
        return None
    try:
        return int(float(s.replace(",", ".")))
    except ValueError:
        return None


def parse_meta_flag(value) -> bool | None:
    if value is None or str(value).strip() == "":
        return None
    s = str(value).strip().casefold()
    if s in {"да", "yes", "y", "1", "+"}:
        return True
    if s in {"нет", "no", "n", "0", "-"}:
        return False
    return None


def load_core_rows():
    wb = load_workbook(XLSX_PATH, data_only=True)
    ws = wb[SHEET_CORE]
    cols = header_map(ws)
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        service = row[cols[HEADER_SERVICE] - 1]
        keyword = row[cols[HEADER_KEYWORD] - 1]
        if not service or not keyword:
            continue
        text_freq = parse_int(row[cols[HEADER_TEXT_FREQ] - 1]) if HEADER_TEXT_FREQ in cols else None
        ws_freq = parse_int(row[cols[HEADER_WS_FREQ] - 1]) if HEADER_WS_FREQ in cols else None
        in_meta = parse_meta_flag(row[cols[HEADER_IN_META] - 1]) if HEADER_IN_META in cols else None
        rows.append(
            {
                "service": str(service).strip(),
                "keyword": str(keyword).strip(),
                "text_freq": text_freq,
                "wordstat_freq": ws_freq,
                "in_meta": in_meta,
            }
        )
    return rows

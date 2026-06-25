#!/usr/bin/env python3
"""Собирает meta keywords из xlsx и записывает в CMS JSON.

Отбор фраз:
  - «В meta» = да  → включить
  - «В meta» = нет → исключить
  - пусто          → включить, если частотность Вордстата >= --min-frequency

Примеры:
  python3 docs/seo/apply_keywords_from_xlsx.py --dry-run
  python3 docs/seo/apply_keywords_from_xlsx.py --apply --min-frequency 10 --max-per-page 12
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from seo_xlsx_common import (
    HEADER_IN_META,
    HEADER_KEYWORD,
    HEADER_SERVICE,
    HEADER_WS_FREQ,
    SERVICE_TO_SLUG,
    SHEET_CMS,
    SHEET_CORE,
    XLSX_PATH,
    load_core_rows,
)

ROOT = Path(__file__).resolve().parents[2]
PAGES_PATH = ROOT / "backend/data/pages.import.json"
FOLDERS_PATH = ROOT / "backend/data/folders.json"
KEYWORDS_MAX_LEN = 1000


def join_keywords(keywords: list[str]) -> str:
    parts: list[str] = []
    for raw in keywords:
        kw = raw.strip()
        if not kw:
            continue
        next_line = f"{', '.join(parts)}, {kw}" if parts else kw
        if len(next_line) > KEYWORDS_MAX_LEN:
            break
        parts.append(kw)
    return ", ".join(parts)


def select_keywords(rows, min_frequency: int, max_per_page: int) -> dict[str, list[str]]:
    by_service: dict[str, list[tuple[int, str]]] = {}

    for row in rows:
        service = row["service"]
        keyword = row["keyword"]
        flag = row["in_meta"]
        freq = row["wordstat_freq"]

        if flag is False:
            continue
        if flag is True:
            by_service.setdefault(service, []).append((freq or 0, keyword))
            continue
        if freq is None:
            continue
        if freq >= min_frequency:
            by_service.setdefault(service, []).append((freq, keyword))

    result: dict[str, list[str]] = {}
    for service, items in by_service.items():
        items.sort(key=lambda x: (-x[0], x[1]))
        seen: set[str] = set()
        picked: list[str] = []
        for _freq, kw in items:
            key = kw.casefold()
            if key in seen:
                continue
            seen.add(key)
            picked.append(kw)
            if len(picked) >= max_per_page:
                break
        result[service] = picked
    return result


def set_page_keywords(pages: list[dict], slug: str, keywords: str) -> bool:
    page = next((p for p in pages if p.get("slug") == slug), None)
    if not page:
        return False
    blocks = page.setdefault("blocks", [])
    block = next((b for b in blocks if b.get("type") == "keywords"), None)
    if block is None:
        blocks.append({"type": "keywords", "order": 1, "data": {"text": keywords}})
    else:
        block.setdefault("data", {})["text"] = keywords
    page["keywords"] = keywords
    return True


def write_cms_sheet(selected: dict[str, list[str]]) -> None:
    wb = load_workbook(XLSX_PATH)
    if SHEET_CMS in wb.sheetnames:
        del wb[SHEET_CMS]
    ws = wb.create_sheet(SHEET_CMS)
    ws.append(["Услуга", "Slug / цель", "Keywords для CMS", "Кол-во"])
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="496DB3")
    for col in range(1, 5):
        c = ws.cell(1, col)
        c.font = header_font
        c.fill = header_fill
    for service, kws in selected.items():
        slug = SERVICE_TO_SLUG.get(service, "")
        line = join_keywords(kws)
        target = "folders.json → catalogization" if slug == "__hub__" else slug
        ws.append([service, target, line, len(kws)])
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 72
    ws.column_dimensions["D"].width = 8
    wb.save(XLSX_PATH)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true", help="Только вывод и лист CMS, без JSON")
    parser.add_argument("--apply", action="store_true", help="Записать в pages.import.json и folders.json")
    parser.add_argument("--min-frequency", type=int, default=1, help="Мин. показов/мес в Вордстате")
    parser.add_argument("--max-per-page", type=int, default=12, help="Макс. фраз в meta keywords")
    args = parser.parse_args()

    if not args.dry_run and not args.apply:
        args.dry_run = True

    rows = load_core_rows()
    selected = select_keywords(rows, args.min_frequency, args.max_per_page)
    write_cms_sheet(selected)

    for service in SERVICE_TO_SLUG:
        kws = selected.get(service, [])
        slug = SERVICE_TO_SLUG[service]
        line = join_keywords(kws)
        print(f"\n=== {service} ===")
        print(f"target: {slug}")
        print(f"keywords ({len(kws)}): {line or '(пусто — заполните Вордстат)'}")

    if not args.apply:
        print(f"\nЛист «{SHEET_CMS}» обновлён. Для записи в CMS: --apply")
        return

    pages = json.loads(PAGES_PATH.read_text(encoding="utf-8"))
    folders = json.loads(FOLDERS_PATH.read_text(encoding="utf-8"))
    changed = 0

    for service, kws in selected.items():
        slug = SERVICE_TO_SLUG.get(service)
        line = join_keywords(kws)
        if not line:
            print(f"Пропуск (нет ключей): {service}")
            continue
        if slug == "__hub__":
            folder = next((f for f in folders if f.get("slug") == "catalogization"), None)
            if folder is None:
                print("Не найден catalogization в folders.json")
                continue
            folder["keywords"] = line
            changed += 1
            continue
        if set_page_keywords(pages, slug, line):
            changed += 1
        else:
            print(f"Страница не найдена: {slug}")

    PAGES_PATH.write_text(json.dumps(pages, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    FOLDERS_PATH.write_text(json.dumps(folders, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"\nЗаписано в CMS: {changed} целей. Импорт на сервер: npm run pages:import (при необходимости).")


if __name__ == "__main__":
    main()

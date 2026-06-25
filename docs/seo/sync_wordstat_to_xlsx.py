#!/usr/bin/env python3
"""Переносит частотности с листа «Вордстат — ввод» на «Семантическое ядро».

После ручного заполнения Вордстата:
  python3 docs/seo/sync_wordstat_to_xlsx.py
"""

from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from openpyxl import load_workbook

from seo_xlsx_common import (
    HEADER_KEYWORD,
    HEADER_SERVICE,
    HEADER_WS_FREQ,
    SHEET_CORE,
    SHEET_WORDSTAT,
    XLSX_PATH,
    header_map,
    parse_int,
)

def main() -> None:
    wb = load_workbook(XLSX_PATH)
    ws_in = wb[SHEET_WORDSTAT]
    freq_by_kw: dict[str, int] = {}
    for row in ws_in.iter_rows(min_row=2, values_only=True):
        kw, freq = row[0], row[1]
        if not kw:
            continue
        parsed = parse_int(freq)
        if parsed is not None:
            freq_by_kw[str(kw).strip()] = parsed

    ws = wb[SHEET_CORE]
    cols = header_map(ws)
    if HEADER_WS_FREQ not in cols:
        raise SystemExit(f"Нет колонки «{HEADER_WS_FREQ}». Пересоздайте структуру xlsx.")

    updated = 0
    for r in range(2, ws.max_row + 1):
        kw = str(ws.cell(r, cols[HEADER_KEYWORD]).value or "").strip()
        if kw in freq_by_kw:
            ws.cell(r, cols[HEADER_WS_FREQ], freq_by_kw[kw])
            updated += 1

    wb.save(XLSX_PATH)
    print(f"Обновлено строк: {updated} (уникальных частот в вводе: {len(freq_by_kw)})")


if __name__ == "__main__":
    main()
